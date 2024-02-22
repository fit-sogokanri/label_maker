# -*- coding:utf-8 -*-
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, A5, landscape, portrait
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import pyqrcode
import openpyxl
import os
import shutil

PDF_FILE_PATH = "./label.pdf"

pdfmetrics.registerFont(TTFont("YuGothR", "C:/Windows/Fonts/YuGothR.ttc"))
pdfmetrics.registerFont(TTFont("YuGothM", "C:/Windows/Fonts/YuGothM.ttc"))

print_size = landscape(A4)
label_count_per_page = 45
left_margin = 5
button_margin = 20

qr_base_url = ""
org_name = ""

labels = []


class label_data:
    org = ""
    manage_code = ""
    date = ""
    link_code = ""
    text1 = ""
    text2 = ""
    text3 = ""

    def __init__(self, org, manage_code, date, link_code, text1, text2="", text3=""):
        if not text2 is None:
            self.text2 = text2.__str__()
        if not text3 is None:
            self.text3 = text3.__str__()

        self.org = org.__str__()
        self.manage_code = manage_code.__str__()
        self.date = date.__str__()
        self.link_code = link_code.__str__()
        self.text1 = text1.__str__()


def main():
    os.makedirs("./qr", exist_ok=True)
    load_setting_file("./setting.xlsx")
    print("--- start ---")
    create_pdf(PDF_FILE_PATH)
    print("--- end ---")
    shutil.rmtree("./qr")


def load_setting_file(file_name):
    wb = openpyxl.load_workbook(filename=file_name, data_only=True)

    global qr_base_url, org_name, print_size, label_count_per_page, left_margin, button_margin
    setting_ws = wb["設定"]
    print_size_param = setting_ws.cell(row=2, column=2).value
    qr_base_url = setting_ws.cell(row=3, column=2).value
    org_name = setting_ws.cell(row=4, column=2).value
    if print_size_param == "A5":
        print_size = portrait(A5)
        label_count_per_page = 18
        left_margin = 7
        button_margin = 20
    if qr_base_url is None:
        print("base url is empty.")
        exit()
    if org_name is None:
        print("organization name is empty.")
        exit()

    label_ws = wb["ラベル"]
    for row in label_ws.rows:
        if row[1].value is None or row[1].value == "":
            break
        if row[1].value == "管理組織":
            continue

        labels.append(
            label_data(
                row[1].value,
                row[2].value,
                row[3].value,
                row[4].value,
                row[5].value,
                row[6].value,
                row[7].value,
            )
        )


def create_pdf(file_path):
    count = 0
    pdf_canvas = canvas.Canvas(file_path, pagesize=print_size)

    pdf_canvas.setTitle("備品ラベル")
    pdf_canvas.setSubject("バージョン1.0.0")

    page_count = (len(labels) + label_count_per_page - 1) // label_count_per_page

    # pageを設定
    for page in range(1, page_count + 1):
        for y in range(left_margin, int(print_size[1] / mm), 22):
            pdf_canvas.setStrokeColorRGB(0.75, 0.75, 0.75)
            pdf_canvas.setLineWidth(0.3 * mm)
            for x in range(button_margin, int(print_size[0] / mm), 52):
                # 横
                pdf_canvas.line((x - 3) * mm, y * mm, (x + 3) * mm, y * mm)
                # 縦
                pdf_canvas.line(x * mm, (y - 3) * mm, x * mm, (y + 3) * mm)

        for y in range(left_margin + 1, int(print_size[1] / mm) - 22, 22):
            pdf_canvas.setStrokeColorRGB(0, 0, 0)
            pdf_canvas.setLineWidth(0.5 * mm)
            for x in range(button_margin + 1, int(print_size[0] / mm) - 52, 52):
                index = label_count_per_page * page - count - 1 + label_count_per_page * (page - 1)

                if count == label_count_per_page * page_count:
                    break
                if index >= len(labels):
                    count += 1
                    continue

                # ライン描画
                pdf_canvas.rect(x * mm, y * mm, 50 * mm, 20 * mm)
                # 下ライン
                pdf_canvas.line(x * mm, (y + 4) * mm, (x + 50) * mm, (y + 4) * mm)
                # 上ライン
                pdf_canvas.line(x * mm, (y + 16) * mm, (x + 35) * mm, (y + 16) * mm)
                # 上ライン 管理組織 縦
                pdf_canvas.line((x + 6) * mm, (y + 16) * mm, (x + 6) * mm, (y + 20) * mm)
                # 上ライン 管理番号 縦
                pdf_canvas.line((x + 21) * mm, (y + 16) * mm, (x + 21) * mm, (y + 20) * mm)
                # 上ライン 取得年月日 縦
                pdf_canvas.line((x + 35) * mm, (y + 15.749) * mm, (x + 35) * mm, (y + 20) * mm)

                # 文字入力
                pdf_canvas.setFont("Helvetica", 2.6 * mm)
                # 取得年月日
                pdf_canvas.drawCentredString((x + 28) * mm, (y + 16.9) * mm, labels[index].date)
                # 管理番号
                pdf_canvas.drawCentredString((x + 13.5) * mm, (y + 16.9) * mm, labels[index].manage_code)
                # 下ラベル
                pdf_canvas.setFont("YuGothM", 3 * mm)
                pdf_canvas.drawCentredString((x + 25) * mm, (y + 0.75) * mm, org_name)
                # 管理組織
                pdf_canvas.setFont("YuGothM", 2.6 * mm)
                pdf_canvas.drawCentredString((x + 3) * mm, (y + 16.9) * mm, labels[index].org)
                pdf_canvas.drawString((x + 0.5) * mm, (y + 12) * mm, labels[index].text1)
                pdf_canvas.drawString((x + 0.5) * mm, (y + 9) * mm, labels[index].text2)
                pdf_canvas.drawString((x + 0.5) * mm, (y + 6) * mm, labels[index].text3)
                qr = pyqrcode.create(f"{qr_base_url}/?c={labels[index].link_code}", version=3, error='M')
                qr.png(f"./qr/qrcode{index}.png", scale=3)
                pdf_canvas.drawImage(f"./qr/qrcode{index}.png", (x + 35.4) * mm, (y + 5) * mm, 14 * mm, 14 * mm)

                count = count + 1

        pdf_canvas.showPage()
    pdf_canvas.save()  # pdfを保存


if __name__ == '__main__':
    main()
